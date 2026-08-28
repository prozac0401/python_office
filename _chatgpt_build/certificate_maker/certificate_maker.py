# -*- coding: utf-8 -*-
"""
Win32 COM 기반 인증서 자동 제작 도구
- template .pptx 내부의 {컬럼명} 변수를 엑셀 첫 번째 시트의 헤더/값으로 치환
- 템플릿 파일명 또는 UI의 출력 파일명 패턴에 있는 {컬럼명}도 치환
- PPTX, PDF, PNG 출력 지원

실행 환경: Windows + Microsoft PowerPoint 설치 필요
"""
from __future__ import annotations

import os
import re
import sys
import shutil
import threading
import queue
import traceback
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter.scrolledtext import ScrolledText

try:
    from openpyxl import load_workbook
except Exception:  # EXE 환경에서 의존성 누락 안내를 위해 런타임에서 처리
    load_workbook = None

APP_TITLE = "인증서 자동 제작 도구"
APP_VERSION = "1.1.0"

# PowerPoint SaveAs constants
PP_SAVE_AS_OPENXML_PRESENTATION = 24  # .pptx
PP_SAVE_AS_PDF = 32                   # .pdf
PP_ALERTS_NONE = 1

TOKEN_RE = re.compile(r"\{([^{}]+)\}")
INVALID_FILENAME_RE = re.compile(r'[\\/:*?"<>|\r\n\t]+')


@dataclass
class JobSettings:
    template_path: Path
    data_path: Path
    output_dir: Path
    filename_pattern: str
    export_pptx: bool
    export_pdf: bool
    export_png: bool
    png_width: Optional[int]
    png_height: Optional[int]
    keep_powerpoint_visible: bool


def resource_path(relative_path: str) -> Path:
    """PyInstaller one-file EXE와 소스 실행 모두에서 리소스 경로를 찾습니다."""
    base = getattr(sys, "_MEIPASS", None)
    if base:
        return Path(base) / relative_path
    return Path(__file__).resolve().parent / relative_path


def get_pptx_aspect_ratio(pptx_path: Path) -> Optional[float]:
    """PPTX 내부 presentation.xml에서 슬라이드 가로/세로 비율을 읽습니다."""
    try:
        with zipfile.ZipFile(pptx_path, "r") as zf:
            xml_bytes = zf.read("ppt/presentation.xml")
        root = ET.fromstring(xml_bytes)
        ns = {"p": "http://schemas.openxmlformats.org/presentationml/2006/main"}
        sld_sz = root.find("p:sldSz", ns)
        if sld_sz is None:
            return None
        cx = int(sld_sz.attrib.get("cx", "0"))
        cy = int(sld_sz.attrib.get("cy", "0"))
        if cx > 0 and cy > 0:
            return cx / cy
    except Exception:
        return None
    return None


def format_cell_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value)


def sanitize_filename(name: str, fallback: str = "certificate") -> str:
    name = INVALID_FILENAME_RE.sub("_", name).strip()
    name = re.sub(r"\s+", " ", name).strip(" .")
    return name or fallback


def replace_tokens(text: str, row_map: Dict[str, str], leave_unknown: bool = True) -> str:
    def repl(match: re.Match) -> str:
        key = match.group(1).strip()
        if key in row_map:
            return row_map[key]
        return match.group(0) if leave_unknown else ""
    return TOKEN_RE.sub(repl, text)


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem, suffix = path.stem, path.suffix
    parent = path.parent
    idx = 2
    while True:
        candidate = parent / f"{stem}_{idx}{suffix}"
        if not candidate.exists():
            return candidate
        idx += 1


def load_excel_rows(xlsx_path: Path) -> Tuple[List[str], List[Dict[str, str]]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl 모듈을 찾을 수 없습니다. requirements.txt 설치 또는 EXE 빌드를 확인하세요.")

    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb.worksheets[0]

    header_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_values:
        raise ValueError("엑셀 첫 번째 시트의 첫 줄에 헤더가 없습니다.")

    headers: List[str] = []
    seen = set()
    for idx, raw in enumerate(header_values, start=1):
        header = format_cell_value(raw).strip()
        if not header:
            header = f"Column{idx}"
        if header in seen:
            raise ValueError(f"중복 헤더가 있습니다: {header}")
        seen.add(header)
        headers.append(header)

    rows: List[Dict[str, str]] = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if excel_row is None or all(v is None for v in excel_row):
            continue
        row_map = {}
        for idx, header in enumerate(headers):
            value = excel_row[idx] if idx < len(excel_row) else None
            row_map[header] = format_cell_value(value)
        rows.append(row_map)

    if not rows:
        raise ValueError("엑셀 데이터 행이 없습니다. 2행부터 데이터를 입력하세요.")
    return headers, rows


def replace_text_range(text_range, row_map: Dict[str, str]) -> None:
    # TextRange.Replace는 서식을 최대한 보존합니다. 실패 시 전체 텍스트 치환으로 fallback합니다.
    for key, value in row_map.items():
        token = "{" + key + "}"
        try:
            text_range.Replace(FindWhat=token, ReplaceWhat=value, MatchCase=False, WholeWords=False)
        except Exception:
            try:
                current = text_range.Text
                if token in current:
                    text_range.Text = current.replace(token, value)
            except Exception:
                pass


def replace_shape(shape, row_map: Dict[str, str]) -> None:
    # 일반 텍스트 상자 / 도형 텍스트
    try:
        if getattr(shape, "HasTextFrame", 0):
            text_frame = shape.TextFrame
            if getattr(text_frame, "HasText", 0):
                replace_text_range(text_frame.TextRange, row_map)
    except Exception:
        pass

    # 표 내부 셀
    try:
        if getattr(shape, "HasTable", 0):
            table = shape.Table
            for r in range(1, table.Rows.Count + 1):
                for c in range(1, table.Columns.Count + 1):
                    try:
                        cell_shape = table.Cell(r, c).Shape
                        if getattr(cell_shape, "HasTextFrame", 0) and getattr(cell_shape.TextFrame, "HasText", 0):
                            replace_text_range(cell_shape.TextFrame.TextRange, row_map)
                    except Exception:
                        pass
    except Exception:
        pass

    # 그룹 도형
    try:
        group_items = shape.GroupItems
        for idx in range(1, group_items.Count + 1):
            replace_shape(group_items.Item(idx), row_map)
    except Exception:
        pass


def replace_presentation(presentation, row_map: Dict[str, str]) -> None:
    for slide_idx in range(1, presentation.Slides.Count + 1):
        slide = presentation.Slides(slide_idx)
        for shape_idx in range(1, slide.Shapes.Count + 1):
            replace_shape(slide.Shapes(shape_idx), row_map)


def calc_png_size(presentation, width: Optional[int], height: Optional[int]) -> Tuple[int, int]:
    slide_w = float(presentation.PageSetup.SlideWidth)
    slide_h = float(presentation.PageSetup.SlideHeight)
    if not width and not height:
        width = 1920
        height = int(round(width * slide_h / slide_w))
    elif width and not height:
        height = int(round(width * slide_h / slide_w))
    elif height and not width:
        width = int(round(height * slide_w / slide_h))
    return int(width), int(height)


def parse_dimension(raw: str, field_name: str) -> Optional[int]:
    raw = raw.strip()
    if not raw:
        return None
    try:
        value = int(raw)
    except ValueError:
        raise ValueError(f"{field_name}은 숫자로 입력하세요.")
    if value < 100 or value > 10000:
        raise ValueError(f"{field_name}은 100~10000 사이로 입력하세요.")
    return value


def export_for_row(ppt_app, settings: JobSettings, row_map: Dict[str, str], row_number: int, log) -> None:
    presentation = None
    try:
        # FileName, ReadOnly, Untitled, WithWindow
        presentation = ppt_app.Presentations.Open(str(settings.template_path), True, False, settings.keep_powerpoint_visible)
        replace_presentation(presentation, row_map)

        raw_base = replace_tokens(settings.filename_pattern, row_map, leave_unknown=False)
        base_name = sanitize_filename(raw_base, fallback=f"certificate_{row_number:04d}")

        if settings.export_pptx:
            pptx_path = unique_path(settings.output_dir / f"{base_name}.pptx")
            presentation.SaveAs(str(pptx_path), PP_SAVE_AS_OPENXML_PRESENTATION)
            log(f"  PPTX 저장: {pptx_path.name}")

        if settings.export_pdf:
            pdf_path = unique_path(settings.output_dir / f"{base_name}.pdf")
            presentation.SaveAs(str(pdf_path), PP_SAVE_AS_PDF)
            log(f"  PDF 저장: {pdf_path.name}")

        if settings.export_png:
            png_w, png_h = calc_png_size(presentation, settings.png_width, settings.png_height)
            slide_count = presentation.Slides.Count
            for slide_idx in range(1, slide_count + 1):
                if slide_count == 1:
                    png_name = f"{base_name}.png"
                else:
                    png_name = f"{base_name}_slide{slide_idx:02d}.png"
                png_path = unique_path(settings.output_dir / png_name)
                presentation.Slides(slide_idx).Export(str(png_path), "PNG", png_w, png_h)
                log(f"  PNG 저장: {png_path.name} ({png_w}x{png_h})")
    finally:
        if presentation is not None:
            try:
                presentation.Close()
            except Exception:
                pass


def run_job(settings: JobSettings, progress_cb, log_cb, done_cb) -> None:
    try:
        headers, rows = load_excel_rows(settings.data_path)
        settings.output_dir.mkdir(parents=True, exist_ok=True)
        log_cb(f"엑셀 로드 완료: {len(rows)}건 / 헤더: {', '.join(headers)}")
        log_cb("PowerPoint COM 시작 중...")

        try:
            import pythoncom
            import win32com.client
        except Exception as e:
            raise RuntimeError("pywin32/win32com을 사용할 수 없습니다. Windows에서 requirements.txt 설치 또는 EXE 빌드를 확인하세요.") from e

        pythoncom.CoInitialize()
        ppt_app = None
        success_count = 0
        error_count = 0
        try:
            ppt_app = win32com.client.DispatchEx("PowerPoint.Application")
            ppt_app.Visible = True if settings.keep_powerpoint_visible else False
            try:
                ppt_app.DisplayAlerts = PP_ALERTS_NONE
            except Exception:
                pass

            total = len(rows)
            for idx, row_map in enumerate(rows, start=1):
                name_hint = row_map.get("이름", f"{idx}행")
                log_cb(f"[{idx}/{total}] 처리 시작: {name_hint}")
                try:
                    export_for_row(ppt_app, settings, row_map, idx, log_cb)
                    success_count += 1
                    log_cb(f"[{idx}/{total}] 완료")
                except Exception as row_error:
                    error_count += 1
                    log_cb(f"[{idx}/{total}] 오류: {row_error}")
                progress_cb(idx, total)
        finally:
            if ppt_app is not None:
                try:
                    ppt_app.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()

        if error_count:
            done_cb(False, f"완료: 성공 {success_count}건, 오류 {error_count}건")
        else:
            done_cb(True, f"완료: {success_count}건 생성")
    except Exception as e:
        log_cb(traceback.format_exc())
        done_cb(False, str(e))


class CertificateMakerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"{APP_TITLE} v{APP_VERSION}")
        self.geometry("900x760")
        self.minsize(840, 700)

        self.ui_queue: queue.Queue = queue.Queue()
        self.worker: Optional[threading.Thread] = None

        self.template_var = tk.StringVar()
        self.data_var = tk.StringVar()
        self.output_var = tk.StringVar(value=str(Path.cwd() / "output"))
        self.filename_var = tk.StringVar(value="인증서_{이름}_{과정명}")
        self.export_pptx_var = tk.BooleanVar(value=True)
        self.export_pdf_var = tk.BooleanVar(value=True)
        self.export_png_var = tk.BooleanVar(value=False)
        # 세로형(A4 계열) 기본값. 슬라이더는 400~5000px 범위입니다.
        self.png_width_var = tk.DoubleVar(value=1240)
        self.png_height_var = tk.DoubleVar(value=1754)
        self.png_lock_ratio_var = tk.BooleanVar(value=True)
        self.png_ratio = 1240 / 1754
        self._syncing_png_size = False
        # PowerPoint 창은 기본적으로 숨김 처리합니다.
        self.visible_var = tk.BooleanVar(value=False)

        self._build_ui()
        self.after(100, self._poll_queue)

    def _build_ui(self) -> None:
        pad = {"padx": 10, "pady": 6}
        root = ttk.Frame(self)
        root.pack(fill="both", expand=True)

        header = ttk.Frame(root)
        header.pack(fill="x", padx=14, pady=(12, 8))
        ttk.Label(header, text="인증서 자동 제작", font=("맑은 고딕", 16, "bold")).pack(anchor="w")
        ttk.Label(
            header,
            text="PPTX 템플릿의 {헤더명} 변수를 엑셀 첫 번째 시트 데이터로 치환하여 PPTX/PDF/PNG 파일을 생성합니다.",
            foreground="#555555",
        ).pack(anchor="w", pady=(4, 0))

        notice = ttk.LabelFrame(root, text="작업 전 유의사항")
        notice.pack(fill="x", padx=14, pady=(0, 8))
        ttk.Label(
            notice,
            text="Windows + Microsoft PowerPoint가 필요합니다. 실행 중에는 템플릿/출력 파일을 열어두지 말고, PowerPoint 창 표시는 기본 해제되어 있습니다. 창 표시를 켠 경우 작업 중 PowerPoint 조작은 피해주세요. 파일명 중복은 자동으로 _2, _3 형식으로 회피합니다.",
            wraplength=820,
            foreground="#7A3E00",
        ).pack(anchor="w", padx=10, pady=7)

        form = ttk.LabelFrame(root, text="입력 파일")
        form.pack(fill="x", padx=14, pady=8)

        self._file_row(form, "템플릿 PPTX", self.template_var, self._browse_template, row=0)
        self._file_row(form, "데이터 XLSX", self.data_var, self._browse_data, row=1)
        self._file_row(form, "출력 폴더", self.output_var, self._browse_output, row=2)

        ttk.Label(form, text="출력 파일명 패턴").grid(row=3, column=0, sticky="w", **pad)
        ttk.Entry(form, textvariable=self.filename_var).grid(row=3, column=1, sticky="ew", **pad)
        ttk.Label(form, text="예: 수료증_{이름}_{과정명}", foreground="#666666").grid(row=3, column=2, sticky="w", **pad)
        form.columnconfigure(1, weight=1)

        options = ttk.LabelFrame(root, text="출력 형식")
        options.pack(fill="x", padx=14, pady=8)
        ttk.Checkbutton(options, text="PPTX", variable=self.export_pptx_var).grid(row=0, column=0, sticky="w", **pad)
        ttk.Checkbutton(options, text="PDF", variable=self.export_pdf_var).grid(row=0, column=1, sticky="w", **pad)
        ttk.Checkbutton(options, text="PNG", variable=self.export_png_var).grid(row=0, column=2, sticky="w", **pad)
        ttk.Checkbutton(
            options, text="PowerPoint 창 표시", variable=self.visible_var
        ).grid(row=0, column=3, sticky="w", padx=14, pady=6)
        ttk.Checkbutton(
            options, text="PNG 비율 고정", variable=self.png_lock_ratio_var, command=self._toggle_png_ratio_lock
        ).grid(row=0, column=4, sticky="w", padx=14, pady=6)

        ttk.Label(options, text="PNG 너비").grid(row=1, column=0, sticky="w", padx=(10, 4), pady=(6, 2))
        self.png_width_scale = ttk.Scale(
            options, from_=400, to=5000, variable=self.png_width_var, command=self._on_png_width_changed
        )
        self.png_width_scale.grid(row=1, column=1, columnspan=3, sticky="ew", padx=4, pady=(6, 2))
        self.png_width_value = ttk.Label(options, text="1240 px", width=10, anchor="e")
        self.png_width_value.grid(row=1, column=4, sticky="e", padx=10, pady=(6, 2))

        ttk.Label(options, text="PNG 높이").grid(row=2, column=0, sticky="w", padx=(10, 4), pady=(2, 6))
        self.png_height_scale = ttk.Scale(
            options, from_=400, to=5000, variable=self.png_height_var, command=self._on_png_height_changed
        )
        self.png_height_scale.grid(row=2, column=1, columnspan=3, sticky="ew", padx=4, pady=(2, 6))
        self.png_height_value = ttk.Label(options, text="1754 px", width=10, anchor="e")
        self.png_height_value.grid(row=2, column=4, sticky="e", padx=10, pady=(2, 6))

        self.png_ratio_label = ttk.Label(
            options, text="비율 고정: 템플릿 슬라이드 비율", foreground="#666666"
        )
        self.png_ratio_label.grid(row=3, column=0, columnspan=5, sticky="w", padx=10, pady=(0, 8))
        for c in range(1, 4):
            options.columnconfigure(c, weight=1)

        actions = ttk.Frame(root)
        actions.pack(fill="x", padx=14, pady=8)
        ttk.Button(actions, text="예제 파일 만들기", command=self._copy_samples).pack(side="left")
        ttk.Button(actions, text="실행 시작", command=self._start_job).pack(side="right")

        progress_frame = ttk.LabelFrame(root, text="진행 상황")
        progress_frame.pack(fill="both", expand=True, padx=14, pady=(8, 14))
        self.progress = ttk.Progressbar(progress_frame, orient="horizontal", mode="determinate")
        self.progress.pack(fill="x", padx=10, pady=(10, 6))
        self.log_text = ScrolledText(progress_frame, height=14, font=("Consolas", 10))
        self.log_text.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self._log("준비 완료. 예제 파일 만들기로 세로형 인증서 템플릿과 데이터 샘플을 바로 테스트할 수 있습니다.")

    def _update_png_labels(self) -> None:
        self.png_width_value.configure(text=f"{int(round(self.png_width_var.get()))} px")
        self.png_height_value.configure(text=f"{int(round(self.png_height_var.get()))} px")

    def _on_png_width_changed(self, _value=None) -> None:
        if self._syncing_png_size:
            return
        self._syncing_png_size = True
        try:
            width = max(400, min(5000, int(round(self.png_width_var.get()))))
            if self.png_lock_ratio_var.get() and self.png_ratio > 0:
                height = int(round(width / self.png_ratio))
                height = max(400, min(5000, height))
                self.png_height_var.set(height)
            self._update_png_labels()
        finally:
            self._syncing_png_size = False

    def _on_png_height_changed(self, _value=None) -> None:
        if self._syncing_png_size:
            return
        self._syncing_png_size = True
        try:
            height = max(400, min(5000, int(round(self.png_height_var.get()))))
            if self.png_lock_ratio_var.get() and self.png_ratio > 0:
                width = int(round(height * self.png_ratio))
                width = max(400, min(5000, width))
                self.png_width_var.set(width)
            self._update_png_labels()
        finally:
            self._syncing_png_size = False

    def _toggle_png_ratio_lock(self) -> None:
        if self.png_lock_ratio_var.get():
            # 고정 활성화 시 현재 너비를 기준으로 템플릿 비율에 맞춰 높이를 즉시 동기화합니다.
            self._on_png_width_changed()
        self._update_png_labels()

    def _apply_template_ratio(self, pptx_path: Path) -> None:
        ratio = get_pptx_aspect_ratio(pptx_path)
        if not ratio or ratio <= 0:
            self.png_ratio_label.configure(text="비율 고정: 템플릿 비율을 읽지 못했습니다. 현재 비율을 유지합니다.")
            return
        self.png_ratio = ratio
        orientation = "세로" if ratio < 1 else ("가로" if ratio > 1 else "정사각")
        self.png_ratio_label.configure(text=f"비율 고정: 템플릿 슬라이드 비율 ({orientation}, {ratio:.4f}:1)")
        if self.png_lock_ratio_var.get():
            self._on_png_width_changed()

    def _file_row(self, parent, label: str, var: tk.StringVar, command, row: int) -> None:
        pad = {"padx": 10, "pady": 6}
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", **pad)
        ttk.Entry(parent, textvariable=var).grid(row=row, column=1, sticky="ew", **pad)
        ttk.Button(parent, text="찾기", command=command).grid(row=row, column=2, sticky="e", **pad)

    def _browse_template(self) -> None:
        path = filedialog.askopenfilename(title="템플릿 PPTX 선택", filetypes=[("PowerPoint", "*.pptx"), ("All files", "*.*")])
        if path:
            self.template_var.set(path)
            stem = Path(path).stem
            if stem:
                self.filename_var.set(stem)
            self._apply_template_ratio(Path(path))

    def _browse_data(self) -> None:
        path = filedialog.askopenfilename(title="데이터 XLSX 선택", filetypes=[("Excel", "*.xlsx"), ("All files", "*.*")])
        if path:
            self.data_var.set(path)

    def _browse_output(self) -> None:
        path = filedialog.askdirectory(title="출력 폴더 선택")
        if path:
            self.output_var.set(path)

    def _copy_samples(self) -> None:
        dest = self.output_var.get().strip()
        if not dest:
            dest = filedialog.askdirectory(title="예제 파일을 복사할 폴더 선택")
            if not dest:
                return
            self.output_var.set(dest)
        dest_path = Path(dest)
        dest_path.mkdir(parents=True, exist_ok=True)

        sample_dir = resource_path("samples")
        sample_files = [
            sample_dir / "인증서_{이름}_{과정명}.pptx",
            sample_dir / "data.xlsx",
        ]
        copied = []
        for src in sample_files:
            if not src.exists():
                messagebox.showerror("예제 파일 없음", f"예제 파일을 찾을 수 없습니다:\n{src}")
                return
            target = unique_path(dest_path / src.name)
            shutil.copy2(src, target)
            copied.append(target)

        self.template_var.set(str(copied[0]))
        self.data_var.set(str(copied[1]))
        self.filename_var.set(copied[0].stem)
        self._apply_template_ratio(copied[0])
        self._log(f"예제 파일 생성 완료: {dest_path}")
        messagebox.showinfo("완료", "세로형 예제 PPTX와 XLSX를 출력 폴더에 만들었습니다.")

    def _validate_settings(self) -> JobSettings:
        template_path = Path(self.template_var.get().strip())
        data_path = Path(self.data_var.get().strip())
        output_dir = Path(self.output_var.get().strip())
        filename_pattern = self.filename_var.get().strip()

        if not template_path.exists() or template_path.suffix.lower() != ".pptx":
            raise ValueError("템플릿 PPTX 파일을 선택하세요.")
        if not data_path.exists() or data_path.suffix.lower() != ".xlsx":
            raise ValueError("데이터 XLSX 파일을 선택하세요.")
        if not filename_pattern:
            filename_pattern = template_path.stem
        if not (self.export_pptx_var.get() or self.export_pdf_var.get() or self.export_png_var.get()):
            raise ValueError("PPTX, PDF, PNG 중 하나 이상을 선택하세요.")

        png_width = int(round(self.png_width_var.get())) if self.export_png_var.get() else None
        png_height = int(round(self.png_height_var.get())) if self.export_png_var.get() else None
        if self.export_png_var.get():
            if not 400 <= png_width <= 5000 or not 400 <= png_height <= 5000:
                raise ValueError("PNG 너비/높이는 400~5000px 범위로 설정하세요.")

        return JobSettings(
            template_path=template_path,
            data_path=data_path,
            output_dir=output_dir,
            filename_pattern=filename_pattern,
            export_pptx=self.export_pptx_var.get(),
            export_pdf=self.export_pdf_var.get(),
            export_png=self.export_png_var.get(),
            png_width=png_width,
            png_height=png_height,
            keep_powerpoint_visible=self.visible_var.get(),
        )

    def _start_job(self) -> None:
        if self.worker and self.worker.is_alive():
            messagebox.showwarning("실행 중", "이미 작업이 실행 중입니다.")
            return
        try:
            settings = self._validate_settings()
        except Exception as e:
            messagebox.showerror("입력 확인", str(e))
            return

        self.progress["value"] = 0
        self.progress["maximum"] = 100
        self._log("=== 작업 시작 ===")
        self._log(f"템플릿: {settings.template_path}")
        self._log(f"데이터: {settings.data_path}")
        self._log(f"출력 폴더: {settings.output_dir}")

        self.worker = threading.Thread(
            target=run_job,
            args=(settings, self._thread_progress, self._thread_log, self._thread_done),
            daemon=True,
        )
        self.worker.start()

    def _thread_log(self, text: str) -> None:
        self.ui_queue.put(("log", text))

    def _thread_progress(self, current: int, total: int) -> None:
        self.ui_queue.put(("progress", current, total))

    def _thread_done(self, ok: bool, message: str) -> None:
        self.ui_queue.put(("done", ok, message))

    def _poll_queue(self) -> None:
        try:
            while True:
                item = self.ui_queue.get_nowait()
                kind = item[0]
                if kind == "log":
                    self._log(item[1])
                elif kind == "progress":
                    _, current, total = item
                    self.progress["maximum"] = total
                    self.progress["value"] = current
                elif kind == "done":
                    _, ok, msg = item
                    self._log(msg)
                    self._log("=== 작업 종료 ===")
                    if ok:
                        messagebox.showinfo("완료", msg)
                    else:
                        messagebox.showwarning("확인 필요", msg)
        except queue.Empty:
            pass
        self.after(100, self._poll_queue)

    def _log(self, text: str) -> None:
        now = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert("end", f"[{now}] {text}\n")
        self.log_text.see("end")


def main() -> None:
    app = CertificateMakerApp()
    app.mainloop()


if __name__ == "__main__":
    main()
